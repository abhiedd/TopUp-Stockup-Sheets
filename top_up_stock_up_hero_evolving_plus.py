import streamlit as st
import pandas as pd
import requests
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
from rembg import remove
import zipfile
import io
import re

st.set_page_config(layout="wide")
st.title("Campaign+Asset Multi-Tab Output (Hero-Evolving Style)")

# 1. Robust Google Sheets tab listing and download
def get_gsheet_tabs(url):
    r = requests.get(url)
    # Find all tab names and gids
    # Regex finds: {"id":123456,"name":"TabName"
    pattern = re.compile(r'"gid":(\d+),.*?"name":"(.*?)"', re.DOTALL)
    found = pattern.findall(r.text)
    # Fallback to classic pattern if empty
    if not found:
        found = re.findall(r'"(.*?)",\[\d+,(?:\d+,){5,6}(\d+),', r.text)
        found = [(gid, name) for name, gid in found]
    tabs = [{"name": name, "gid": gid} for gid, name in found]
    return tabs

def download_tab_csv(sheet_id, gid):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    r = requests.get(url)
    r.raise_for_status()
    return pd.read_csv(StringIO(r.content.decode('utf-8')), header=0)

def get_sheet_id(url):
    m = re.search(r"/d/([a-zA-Z0-9-_]+)", url)
    if m:
        return m.group(1)
    return None

def robust_read_csv(csv_file):
    try:
        df = pd.read_csv(csv_file)
    except Exception:
        csv_file.seek(0)
        df = pd.read_csv(csv_file, encoding="latin1")
    df.columns = [str(c).strip() for c in df.columns]
    return df

def make_img_map(product_df):
    img_map = {}
    for _, row in product_df.iterrows():
        pid = str(row['MB_id']).strip()
        img_src = str(row['image_src']).strip()
        if pid and img_src and pid.lower() != 'nan' and img_src.lower() != 'nan':
            img_map[pid] = f"https://file.milkbasket.com/products/{img_src}"
    return img_map

def fix_pid(pid):
    if pd.isna(pid) or str(pid).strip().lower() == "nan":
        return ""
    try:
        pstr = str(int(float(pid))).strip()
        return pstr
    except Exception:
        return str(pid).strip()

def process_df(df, hub, img_map=None):
    df["Campaign Names"] = df["Campaign Names"].ffill()
    df["Asset"] = df["Asset"].ffill()
    rows = []
    for _, row in df.iterrows():
        asset_val = str(row.get("Asset", "")).strip().lower()
        if asset_val in ["atc", "atc background"]:
            continue
        pid1 = fix_pid(row.get("PID1", ""))
        pid2 = fix_pid(row.get("PID2", ""))
        if not pid1 and not pid2:
            continue
        fg = row.get("Grid Details", "")
        key = (str(row.get("Campaign Names", "")).strip(), str(row.get("Asset", "")).strip())
        rows.append({
            "Hub": hub,
            "Campaign Name": key[0],
            "Asset": key[1],
            "Focus Grid": fg,
            "PID1": pid1,
            "PID2": pid2,
            "Img1": img_map.get(pid1, "") if img_map else "",
            "Img2": img_map.get(pid2, "") if img_map else "",
        })
    return rows

def generate_tabs(all_rows):
    tabs = {}
    for r in all_rows:
        tabname = f"{r['Campaign Name']} | {r['Asset']}"
        if tabname not in tabs:
            tabs[tabname] = []
        tabs[tabname].append({
            "Hub": r["Hub"],
            "Focus Grid": r["Focus Grid"],
            "PID1": r["PID1"],
            "PID2": r["PID2"],
            "Img1": r["Img1"],
            "Img2": r["Img2"]
        })
    return tabs

def clean_sheet_name(name):
    return re.sub(r'[\[\]\*:/\\?]', '', str(name)).strip()[:31]

def excel_export(tabs, all_pids_tab):
    output = BytesIO()
    wb = Workbook()
    for tname, rows in tabs.items():
        ws = wb.create_sheet(title=clean_sheet_name(tname))
        df = pd.DataFrame(rows)
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        bold_font = Font(bold=True)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = yellow_fill
                    cell.font = bold_font
    # All_PIDs tab (always last)
    ws2 = wb.create_sheet("All_PIDs")
    pid_df = pd.DataFrame(all_pids_tab)
    for r_idx, row in enumerate(dataframe_to_rows(pid_df, index=False, header=True), 1):
        ws2.append(row)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output)
    output.seek(0)
    return output

def get_all_unique_pids(all_rows, img_map):
    pid_set = set()
    for r in all_rows:
        for pid_col in ["PID1", "PID2"]:
            pid = r[pid_col]
            if pid:
                pid_set.add(pid)
    all_pids = sorted(pid_set, key=lambda x: int(x) if x.isdigit() else x)
    rows = []
    for pid in all_pids:
        img_link = img_map.get(pid, "")
        rows.append({"PID": pid, "Img Link": img_link})
    return rows

def has_transparency(img_bytes):
    try:
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            alpha = img.getchannel("A") if "A" in img.getbands() else None
            if alpha and alpha.getextrema()[0] < 255:
                return True
        return False
    except Exception:
        return False

# --- UI ---
input_method = st.radio("Choose input method", ["Paste Google Sheet link", "Upload Excel file"])
sheet_url = None
uploaded_file = None

if input_method == "Paste Google Sheet link":
    sheet_url = st.text_input("Paste your (public) Google Sheet URL")
elif input_method == "Upload Excel file":
    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

product_csv = st.file_uploader("Upload product CSV (with 'MB_id' and 'image_src' columns)", type=["csv"])

img_map = {}
if product_csv:
    product_df = robust_read_csv(product_csv)
    if not {"MB_id", "image_src"}.issubset(product_df.columns):
        st.error("CSV must have columns 'MB_id' and 'image_src'.")
        st.stop()
    product_df["MB_id"] = product_df["MB_id"].apply(fix_pid)
    img_map = make_img_map(product_df)

if sheet_url or uploaded_file:
    all_rows = []
    if sheet_url:
        sheet_id = get_sheet_id(sheet_url)
        if not sheet_id:
            st.error("Invalid Google Sheet URL.")
        else:
            try:
                # PATCH: Use new tab fetch logic
                all_tabs = get_gsheet_tabs(sheet_url)
                for tab in all_tabs:
                    tabname = tab["name"]
                    gid = tab["gid"]
                    try:
                        df = download_tab_csv(sheet_id, gid)
                    except Exception:
                        continue
                    # Only process if it has all key columns
                    if not all(col in df.columns for col in ["Campaign Names", "Asset", "Grid Details", "PID1", "PID2"]):
                        continue
                    rows = process_df(df, tabname, img_map)
                    all_rows.extend(rows)
            except Exception as e:
                st.error(f"Error reading or processing Google Sheet: {e}")
    else:
        xls = pd.ExcelFile(uploaded_file)
        for tab in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=tab, header=0)
            if not all(col in df.columns for col in ["Campaign Names", "Asset", "Grid Details", "PID1", "PID2"]):
                continue
            rows = process_df(df, tab, img_map)
            all_rows.extend(rows)

    if all_rows:
        tabs = generate_tabs(all_rows)
        all_pids_tab = get_all_unique_pids(all_rows, img_map)
        tab_names = sorted(tabs.keys())
        selected_tab = st.selectbox("Select output tab for preview/download:", tab_names)
        preview_df = pd.DataFrame(tabs[selected_tab])
        st.dataframe(preview_df)
        # Excel export
        output = excel_export(tabs, all_pids_tab)
        st.success("✅ Excel file with multi-campaign tabs + All_PIDs ready!")
        st.download_button(
            "📥 Download Multi-Tab Excel",
            output,
            file_name="Campaign_Assetwise_MultiTab_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        with st.expander("Preview All_PIDs tab"):
            st.dataframe(pd.DataFrame(all_pids_tab))

        # --- Download all unique images from All_PIDs ---
        st.markdown("## Download/Process All Unique Images (All_PIDs tab)")
        if st.button("Download ALL images.zip"):
            all_img_rows = [r for r in all_pids_tab if r.get("Img Link")]
            progress = st.progress(0)
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zipf:
                for idx, item in enumerate(all_img_rows):
                    pid = item["PID"]
                    img_url = item["Img Link"]
                    orig_filename = f"{pid}.png"
                    try:
                        r = requests.get(img_url, timeout=10)
                        if r.status_code == 200:
                            img = Image.open(io.BytesIO(r.content)).convert("RGBA")
                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            zipf.writestr(orig_filename, img_byte_arr.getvalue())
                    except Exception:
                        continue
                    progress.progress((idx + 1) / len(all_img_rows))
            zip_buffer.seek(0)
            st.success("All unique images collected! Download ready below.")
            st.download_button(
                label="Download All_PIDs_Images.zip",
                data=zip_buffer,
                file_name="All_PIDs_Images.zip",
                mime="application/zip"
            )

        # --- Optional: Download rembg zip for all unique images ---
        if st.button("Download rembg All images.zip (background removed)"):
            all_img_rows = [r for r in all_pids_tab if r.get("Img Link")]
            progress = st.progress(0)
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zipf:
                for idx, item in enumerate(all_img_rows):
                    pid = item["PID"]
                    img_url = item["Img Link"]
                    orig_filename = f"{pid}.png"
                    try:
                        r = requests.get(img_url, timeout=10)
                        if r.status_code == 200:
                            img = Image.open(io.BytesIO(r.content)).convert("RGBA")
                            if has_transparency(r.content):
                                img_byte_arr = io.BytesIO()
                                img.save(img_byte_arr, format='PNG')
                                zipf.writestr(orig_filename, img_byte_arr.getvalue())
                            else:
                                img = img.resize((650,650))
                                output_image = remove(img)
                                img_byte_arr = io.BytesIO()
                                output_image.save(img_byte_arr, format='PNG')
                                zipf.writestr(orig_filename, img_byte_arr.getvalue())
                    except Exception:
                        continue
                    progress.progress((idx + 1) / len(all_img_rows))
            zip_buffer.seek(0)
            st.success("rembg images ready! Download below.")
            st.download_button(
                label="Download rembg_All_PIDs_Images.zip",
                data=zip_buffer,
                file_name="rembg_All_PIDs_Images.zip",
                mime="application/zip"
            )
    else:
        st.warning("No product data found after filtering.")

